# -*- coding: utf-8 -*-
"""
🥛 NUTRI - Lácteos San Antonio C.A.
Dashboard de Control de Inventario - Sistema de Gestión Logística
Desarrollado por: GEM - Analítica de Datos Logísticos
"""

import streamlit as st
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import plotly.express as px
import plotly.graph_objects as go
from io import BytesIO
import os
import json
import base64

# ============================================
# CONFIGURACIÓN DE LA PÁGINA
# ============================================
st.set_page_config(
    page_title="NUTRI - Control de Inventario",
    page_icon="🥛",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ============================================
# COLORES NUTRI (Azul y Verde)
# ============================================
NUTRI_AZUL = "#0077B6"
NUTRI_AZUL_CLARO = "#00B4D8"
NUTRI_VERDE = "#7CB342"
NUTRI_VERDE_CLARO = "#9CCC65"

# ============================================
# ESTILOS CSS PERSONALIZADOS - TEMA NUTRI
# ============================================
st.markdown(f"""
<style>
    /* Header principal con colores Nutri */
    .main-header {{
        font-size: 2.2rem;
        color: white;
        text-align: center;
        padding: 1.5rem;
        background: linear-gradient(135deg, {NUTRI_AZUL} 0%, {NUTRI_AZUL_CLARO} 50%, {NUTRI_VERDE} 100%);
        border-radius: 15px;
        margin-bottom: 2rem;
        box-shadow: 0 4px 15px rgba(0,119,182,0.3);
    }}
    
    /* Cards de métricas */
    .metric-card {{
        background-color: #f8f9fa;
        padding: 1rem;
        border-radius: 10px;
        border-left: 5px solid {NUTRI_AZUL};
    }}
    
    /* Estados con colores */
    .status-libre {{ background-color: #C8E6C9; padding: 5px 10px; border-radius: 5px; }}
    .status-retenida {{ background-color: #FFCDD2; padding: 5px 10px; border-radius: 5px; }}
    .status-produccion {{ background-color: #FFF9C4; padding: 5px 10px; border-radius: 5px; }}
    
    /* Tabs personalizados */
    .stTabs [data-baseweb="tab-list"] {{ 
        gap: 8px; 
        background-color: transparent;
    }}
    .stTabs [data-baseweb="tab"] {{
        height: 50px;
        padding: 10px 20px;
        background: linear-gradient(180deg, {NUTRI_AZUL_CLARO}22 0%, {NUTRI_AZUL}22 100%);
        border-radius: 10px 10px 0 0;
        border: 2px solid {NUTRI_AZUL}44;
    }}
    .stTabs [data-baseweb="tab"]:hover {{
        background: linear-gradient(180deg, {NUTRI_VERDE}33 0%, {NUTRI_AZUL}33 100%);
    }}
    
    /* Métricas */
    div[data-testid="stMetricValue"] {{ 
        font-size: 1.8rem; 
        color: {NUTRI_AZUL};
    }}
    
    /* Botones primarios */
    .stButton>button[kind="primary"] {{
        background: linear-gradient(90deg, {NUTRI_AZUL} 0%, {NUTRI_VERDE} 100%);
        border: none;
    }}
    
    /* Sidebar con FONDO BLANCO y texto oscuro */
    section[data-testid="stSidebar"] {{
        background-color: #FFFFFF !important;
        background-image: none !important;
    }}
    
    section[data-testid="stSidebar"] > div {{
        background-color: #FFFFFF !important;
    }}
    
    section[data-testid="stSidebar"] .stMarkdown,
    section[data-testid="stSidebar"] .stMarkdown p,
    section[data-testid="stSidebar"] .stMarkdown span,
    section[data-testid="stSidebar"] .stMarkdown li,
    section[data-testid="stSidebar"] p, 
    section[data-testid="stSidebar"] span,
    section[data-testid="stSidebar"] li,
    section[data-testid="stSidebar"] label {{
        color: #333333 !important;
    }}
    
    section[data-testid="stSidebar"] h1,
    section[data-testid="stSidebar"] h2,
    section[data-testid="stSidebar"] h3 {{
        color: {NUTRI_AZUL} !important;
    }}
    
    section[data-testid="stSidebar"] .stMetric label {{
        color: #333333 !important;
    }}
    
    section[data-testid="stSidebar"] [data-testid="stMetricValue"] {{
        color: {NUTRI_AZUL} !important;
    }}
    
    /* Sección info */
    .seccion-info {{
        background: linear-gradient(90deg, {NUTRI_AZUL}15 0%, {NUTRI_VERDE}15 100%);
        padding: 15px;
        border-radius: 10px;
        border-left: 4px solid {NUTRI_VERDE};
        margin: 10px 0;
    }}
</style>
""", unsafe_allow_html=True)

# ============================================
# CATÁLOGO DE PRODUCTOS CON CÓDIGO SAP
# ============================================
PRODUCTOS_SAP = {
    60000002: "LECHE SEMIDESCREMADA UHT POLIETILENO 1L",
    60000026: "AVENA CON MARACUYA UHT TETRA SQUARE 1L",
    60000027: "AVENA CON MARACUYA UHT SLIM 200ML",
    60000028: "AVENA CON LECHE UHT SLIM 200ML",
    60000029: "AVENA CON LECHE UHT TETRA SQUARE 1L",
    60000030: "CREMA DE LECHE UHT TETRAFINO 1/2LT",
    60000031: "CREMA DE LECHE UHT TETRAFINO 100ML",
    60000032: "CREMA DE LECHE UHT POLIETILENO 200ML",
    60000033: "CREMA DE LECHE UHT TETRAFINO 900ML",
    60000035: "LECHE ENTERA UHT POLIETILENO 1L",
    60000036: "LECHE ENTERA UHT SLIM 200ML",
    60000037: "LECHE ENTERA UHT TETRAFINO 1/2 L",
    60000038: "LECHE ENTERA UHT TETRAFINO 1L",
    60000039: "LECHE SEMIDESCREMADA UHT TETRAFINO 1L",
    60000041: "LECHE DESLACTOSADA UHT TETRA SQUARE 1L",
    60000042: "LECHE DESLACTOSADA UHT TETRAFINO 1L",
    60000043: "LECHE DESCREMADA UHT TETRA SQUARE 1L",
    60000044: "LECHE ENTERA UHT TETRA SQUARE 1L",
    60000045: "LECHE SEMIDESCREMADA UHT SQUARE 1L",
    60000047: "LECHE SEMIDESCREMADA UHT TETRAFINO 1/2L",
    60000048: "LECHE DESCREMADA UHT TETRAFINO 1L",
    60000049: "LECHE ENTERA UHT TETRAFINO 900ML",
    60000050: "LECHE SEMIDESCREMADA UHT TETRAFINO 900ML",
    60000051: "LECHE DESCREMADA UHT TETRAFINO 900ML",
    60000052: "LECHE DESLACTOSADA UHT TETRAFINO 900ML",
    60000053: "LECHE ENTERA UHT TETRA SQUARE 500ML",
    60000054: "LECHE SEMIDESCREMADA UHT SQUARE 500ML",
    60000055: "LECHE DESCREMADA UHT TETRA SQUARE 500ML",
    60000056: "LECHE DESLACTOSADA UHT SQUARE 500ML",
    60000057: "LECHE ENTERA UHT POLIETILENO 900ML",
    60000058: "LECHE SEMIDESCREMADA UHT POL 900ML",
    60000113: "LECHE FRESA UHT SLIM 200ML",
    60000114: "LECHE CHOCOLATE UHT LEAF 200 ML",
    60000115: "LECHE ENTERA UHT POLIETILENO 200ML",
    60000120: "4 PACK LECHE ENTERA UHT POLIETILENO 1L",
    60000121: "4 PACK LECHE SEMI UHT POLIETILENO 1L",
    60000122: "3 PACK LECHE DESLACTOSADA SQUARE 1L",
    60000123: "3 PACK LECHE DESCREMADA SQUARE 1L",
    60000124: "4 PACK LECHE ENTERA SQUARE 1L",
    60000125: "4 PACK LECHE SEMIDESCREMADA SQUARE 1L",
    60000130: "6 PACK AVENA CON LECHE SL 200ml",
    60000131: "6 PACK LECHE SABORIZADA FRESA SQ 200ml",
    60000132: "6 PACK LECHE SABOR CHOCOLATE SQ 200ml",
    60000138: "LECHE CHOCOLATE NUTRI UHT SQUARE 1 LT",
    60000172: "AVENA CON NARANJILLA UHT TETRA SQUARE 1L",
    60000173: "AVENA CON NARANJILLA UHT SLIM 200ML",
    60000180: "NUTRI NECTAR NARANJA UHT SQUARE 1L",
    60000182: "NUTRI NECTAR DURAZNO UHT SQUARE 1L",
    60000183: "NUTRI NECTAR NARANJA UHT SLIM 200ML",
    60000184: "NUTRI NECTAR DURAZNO UHT SLIM 200ML",
    60000185: "LECHE DESLACTOSADA UHT POL 900ML",
    60000186: "6 PACK AVENA CON NARANJILLA SQ 200ML",
    60000233: "CREMA DE LECHE SQUARE 1L",
    60000296: "NUTRI AVENA CON MARACUYA UHT POLI 900 ML",
    60000303: "4PACK LECHE UHT DESLACTOSADA POLI 900ML",
    60000312: "LECHE ENTERA UHT POLIETILENO 400ML",
    60000408: "5PACK LECHE ENTERA UHT POLI 900ML",
    60000409: "5PACK LECHE SEMIDESCREMADA UHT POL 900ML",
    60000410: "5PACK LECHE DESLACTOSADA UHT POL 900ML",
    60000539: "NUTRI MEZCLA 3 LECHES UHT SQUARE 1L",
    60000540: "NUTRI LECHE EVAPORADA UHT SQUARE 500ML",
    60000569: "LECHE ENTERA EN POLVO FORTIFICADA 100 G",
    60000570: "LECHE ENTERA EN POLVO FORTIFICADA 200G",
    60000571: "LECHE ENTERA EN POLVO FORTIFICADA 400G",
}

# Lista para selectbox (código - descripción)
PRODUCTOS_LISTA = [f"{codigo} - {desc}" for codigo, desc in sorted(PRODUCTOS_SAP.items(), key=lambda x: x[1])]

ESTADOS = ["LIBRE UTILIZACIÓN", "RETENIDA", "PRODUCCIÓN", "NOVEDADES"]

# Secciones de bodega (filas)
SECCIONES_BODEGA = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J"]
FILAS_BODEGA = list(range(1, 21))  # Filas 1 a 20

# Días de vida útil por tipo de producto
VIDA_UTIL_DIAS = {
    "AVENA": 180,
    "LECHE ENTERA UHT": 180,
    "LECHE SEMIDESCREMADA": 180,
    "LECHE DESLACTOSADA": 180,
    "LECHE DESCREMADA": 180,
    "LECHE CHOCOLATE": 180,
    "LECHE FRESA": 180,
    "LECHE EN POLVO": 365,
    "CREMA": 180,
    "NECTAR": 180,
    "PACK": 180,
    "EVAPORADA": 365,
    "3 LECHES": 365,
    "DEFAULT": 180
}

# Título del sistema (sin logo externo por ahora)
TITULO_SISTEMA = "🥛 NUTRI - Lácteos San Antonio C.A."

# Archivo para persistencia de datos
DATA_FILE = "inventario_lacteos.json"
DATA_PATH = os.path.join(os.path.dirname(__file__), DATA_FILE)

# ============================================
# FUNCIONES DE UTILIDAD
# ============================================
def get_vida_util_dias(producto):
    """Obtiene los días de vida útil según el tipo de producto"""
    producto_upper = producto.upper()
    for key, dias in VIDA_UTIL_DIAS.items():
        if key in producto_upper:
            return dias
    return VIDA_UTIL_DIAS["DEFAULT"]

def calcular_porcentaje_vida_util(fecha_elab, fecha_cad):
    """Calcula el porcentaje de vida útil restante"""
    try:
        hoy = datetime.now().date()
        
        if isinstance(fecha_elab, str):
            fecha_elab = datetime.strptime(fecha_elab, "%Y-%m-%d").date()
        elif isinstance(fecha_elab, datetime):
            fecha_elab = fecha_elab.date()
            
        if isinstance(fecha_cad, str):
            fecha_cad = datetime.strptime(fecha_cad, "%Y-%m-%d").date()
        elif isinstance(fecha_cad, datetime):
            fecha_cad = fecha_cad.date()
        
        vida_total = (fecha_cad - fecha_elab).days
        vida_restante = (fecha_cad - hoy).days
        
        if vida_total <= 0:
            return 0
        return max(0, min(100, (vida_restante / vida_total) * 100))
    except:
        return None

def semaforo_vida_util(porcentaje):
    """Retorna el emoji del semáforo según el porcentaje"""
    if porcentaje is None:
        return "⚪"
    if porcentaje < 30:
        return "🔴"
    elif porcentaje < 50:
        return "🟡"
    else:
        return "🟢"

def color_vida_util_bg(val):
    """Colorea las celdas según el porcentaje de vida útil"""
    if pd.isna(val) or val is None:
        return ''
    if val < 30:
        return 'background-color: #FFCDD2'
    elif val < 50:
        return 'background-color: #FFF9C4'
    else:
        return 'background-color: #C8E6C9'

def color_estado_bg(val):
    """Colorea las celdas según el estado"""
    colores = {
        "LIBRE UTILIZACIÓN": 'background-color: #C8E6C9',
        "RETENIDA": 'background-color: #FFCDD2',
        "PRODUCCIÓN": 'background-color: #FFF9C4',
        "NOVEDADES": 'background-color: #FFE0B2'
    }
    return colores.get(val, '')

# ============================================
# PERSISTENCIA DE DATOS
# ============================================
def guardar_datos():
    """Guarda los datos en un archivo JSON"""
    try:
        datos = {
            'inventario': st.session_state.inventario.to_dict('records'),
            'ultimo_id': st.session_state.ultimo_id
        }
        with open(DATA_PATH, 'w', encoding='utf-8') as f:
            json.dump(datos, f, ensure_ascii=False, indent=2, default=str)
    except Exception as e:
        st.error(f"Error al guardar: {e}")

def cargar_datos():
    """Carga los datos desde el archivo JSON"""
    try:
        if os.path.exists(DATA_PATH):
            with open(DATA_PATH, 'r', encoding='utf-8') as f:
                datos = json.load(f)
                return pd.DataFrame(datos['inventario']), datos['ultimo_id']
    except Exception as e:
        st.warning(f"No se pudieron cargar datos previos: {e}")
    return None, 0

def inicializar_datos():
    """Inicializa el DataFrame en session_state"""
    if 'inventario' not in st.session_state:
        df_cargado, ultimo_id = cargar_datos()
        if df_cargado is not None and not df_cargado.empty:
            st.session_state.inventario = df_cargado
            st.session_state.ultimo_id = ultimo_id
        else:
            st.session_state.inventario = pd.DataFrame(columns=[
                'ID', 'FECHA_REGISTRO', 'ESTADO', 'SECCION', 'FILA', 'CODIGO_SAP', 
                'DESCRIPCION', 'LOTE', 'FECHA_ELABORACION', 'FECHA_CADUCIDAD', 
                'PALLETS', 'CAJAS', 'UNIDADES', 'OBSERVACIONES', 'VIDA_UTIL_%', 'SEMAFORO'
            ])
            st.session_state.ultimo_id = 0

def agregar_registro(estado, seccion, fila, codigo_sap, descripcion, lote, fecha_elab, fecha_cad, pallets, cajas, unidades, observaciones):
    """Agrega un nuevo registro al inventario"""
    st.session_state.ultimo_id += 1
    
    vida_util = calcular_porcentaje_vida_util(fecha_elab, fecha_cad)
    semaforo = semaforo_vida_util(vida_util)
    
    nuevo_registro = pd.DataFrame([{
        'ID': st.session_state.ultimo_id,
        'FECHA_REGISTRO': datetime.now().strftime("%Y-%m-%d %H:%M"),
        'ESTADO': estado,
        'SECCION': seccion,
        'FILA': fila,
        'CODIGO_SAP': codigo_sap,
        'DESCRIPCION': descripcion,
        'LOTE': lote,
        'FECHA_ELABORACION': str(fecha_elab),
        'FECHA_CADUCIDAD': str(fecha_cad),
        'PALLETS': pallets,
        'CAJAS': cajas,
        'UNIDADES': unidades,
        'OBSERVACIONES': observaciones,
        'VIDA_UTIL_%': vida_util,
        'SEMAFORO': semaforo
    }])
    
    st.session_state.inventario = pd.concat([st.session_state.inventario, nuevo_registro], ignore_index=True)
    guardar_datos()

def eliminar_registro(id_registro):
    """Elimina un registro del inventario"""
    st.session_state.inventario = st.session_state.inventario[
        st.session_state.inventario['ID'] != id_registro
    ]
    guardar_datos()

def exportar_excel():
    """Exporta los datos a un archivo Excel con formato según estructura CONTEO"""
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    
    output = BytesIO()
    
    # Colores para cada sección
    COLOR_LIBRE = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Verde claro
    COLOR_LIBRE_HEADER = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Amarillo
    COLOR_RETENIDA = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")  # Rojo
    COLOR_PRODUCCION = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")  # Azul claro
    COLOR_NOVEDADES = PatternFill(start_color="DDA0DD", end_color="DDA0DD", fill_type="solid")  # Morado
    
    header_font = Font(bold=True, size=10)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Columnas del reporte según estructura CONTEO
        columnas_reporte = [
            'SECCION', 'FILA', 'CODIGO_SAP', 'DESCRIPCION', 'LOTE', 
            'PALLETS', 'CAJAS', 'UNIDADES', 'FECHA_ELABORACION', 
            'FECHA_CADUCIDAD', 'VIDA_UTIL_%', 'SEMAFORO', 'OBSERVACIONES'
        ]
        
        # Verificar qué columnas existen
        columnas_existentes = [c for c in columnas_reporte if c in st.session_state.inventario.columns]
        
        # ========== HOJA: LIBRE UTILIZACIÓN ==========
        df_libre = st.session_state.inventario[
            st.session_state.inventario['ESTADO'] == 'LIBRE UTILIZACIÓN'
        ][columnas_existentes].copy() if not st.session_state.inventario.empty else pd.DataFrame(columns=columnas_existentes)
        
        df_libre.to_excel(writer, sheet_name='LIBRE_UTILIZACION', index=False, startrow=1)
        ws_libre = writer.sheets['LIBRE_UTILIZACION']
        ws_libre.merge_cells('A1:M1')
        ws_libre['A1'] = 'LIBRE UTILIZACIÓN'
        ws_libre['A1'].fill = COLOR_LIBRE
        ws_libre['A1'].font = Font(bold=True, size=14)
        ws_libre['A1'].alignment = Alignment(horizontal='center')
        
        # Aplicar colores a encabezados
        for col in range(1, len(columnas_existentes) + 1):
            cell = ws_libre.cell(row=2, column=col)
            cell.fill = COLOR_LIBRE_HEADER
            cell.font = header_font
            cell.border = border
        
        # ========== HOJA: RETENIDA ==========
        df_retenida = st.session_state.inventario[
            st.session_state.inventario['ESTADO'] == 'RETENIDA'
        ][columnas_existentes].copy() if not st.session_state.inventario.empty else pd.DataFrame(columns=columnas_existentes)
        
        df_retenida.to_excel(writer, sheet_name='RETENIDA', index=False, startrow=1)
        ws_retenida = writer.sheets['RETENIDA']
        ws_retenida.merge_cells('A1:M1')
        ws_retenida['A1'] = 'RETENIDA POR FALTA DE LIBERACIÓN'
        ws_retenida['A1'].fill = COLOR_RETENIDA
        ws_retenida['A1'].font = Font(bold=True, size=14, color='FFFFFF')
        ws_retenida['A1'].alignment = Alignment(horizontal='center')
        
        for col in range(1, len(columnas_existentes) + 1):
            cell = ws_retenida.cell(row=2, column=col)
            cell.fill = COLOR_RETENIDA
            cell.font = Font(bold=True, color='FFFFFF')
            cell.border = border
        
        # ========== HOJA: PRODUCCIÓN ==========
        df_produccion = st.session_state.inventario[
            st.session_state.inventario['ESTADO'] == 'PRODUCCIÓN'
        ][columnas_existentes].copy() if not st.session_state.inventario.empty else pd.DataFrame(columns=columnas_existentes)
        
        df_produccion.to_excel(writer, sheet_name='PRODUCCION', index=False, startrow=1)
        ws_produccion = writer.sheets['PRODUCCION']
        ws_produccion.merge_cells('A1:M1')
        ws_produccion['A1'] = 'PRODUCCIÓN'
        ws_produccion['A1'].fill = COLOR_PRODUCCION
        ws_produccion['A1'].font = Font(bold=True, size=14)
        ws_produccion['A1'].alignment = Alignment(horizontal='center')
        
        for col in range(1, len(columnas_existentes) + 1):
            cell = ws_produccion.cell(row=2, column=col)
            cell.fill = COLOR_PRODUCCION
            cell.font = header_font
            cell.border = border
        
        # ========== HOJA: NOVEDADES ==========
        df_novedades = st.session_state.inventario[
            st.session_state.inventario['ESTADO'] == 'NOVEDADES'
        ][columnas_existentes].copy() if not st.session_state.inventario.empty else pd.DataFrame(columns=columnas_existentes)
        
        if not df_novedades.empty or True:  # Siempre crear la hoja
            df_novedades.to_excel(writer, sheet_name='NOVEDADES', index=False, startrow=1)
            ws_novedades = writer.sheets['NOVEDADES']
            ws_novedades.merge_cells('A1:M1')
            ws_novedades['A1'] = 'NOVEDADES'
            ws_novedades['A1'].fill = COLOR_NOVEDADES
            ws_novedades['A1'].font = Font(bold=True, size=14)
            ws_novedades['A1'].alignment = Alignment(horizontal='center')
            
            for col in range(1, len(columnas_existentes) + 1):
                cell = ws_novedades.cell(row=2, column=col)
                cell.fill = COLOR_NOVEDADES
                cell.font = header_font
                cell.border = border
        
        # ========== HOJA: RESUMEN EJECUTIVO ==========
        if not st.session_state.inventario.empty:
            resumen_data = []
            for estado in ESTADOS:
                df_est = st.session_state.inventario[st.session_state.inventario['ESTADO'] == estado]
                resumen_data.append({
                    'ESTADO': estado,
                    'TOTAL_REGISTROS': len(df_est),
                    'TOTAL_PALLETS': df_est['PALLETS'].sum() if not df_est.empty else 0,
                    'TOTAL_CAJAS': df_est['CAJAS'].sum() if not df_est.empty else 0,
                    'TOTAL_UNIDADES': df_est['UNIDADES'].sum() if not df_est.empty else 0
                })
            
            df_resumen = pd.DataFrame(resumen_data)
            df_resumen.to_excel(writer, sheet_name='RESUMEN', index=False)
            
            ws_resumen = writer.sheets['RESUMEN']
            for col in range(1, 6):
                cell = ws_resumen.cell(row=1, column=col)
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.font = Font(bold=True, color='FFFFFF')
                cell.border = border
    
    return output.getvalue()

# ============================================
# INICIALIZAR DATOS
# ============================================
inicializar_datos()

# ============================================
# INTERFAZ PRINCIPAL
# ============================================
st.markdown('<div class="main-header">🥛 NUTRI - Sistema de Control de Inventario<br><small>Lácteos San Antonio C.A.</small></div>', unsafe_allow_html=True)

# ============================================
# PESTAÑAS PRINCIPALES
# ============================================
tab1, tab2, tab3, tab4 = st.tabs(["📝 Ingreso de Datos", "📋 Inventario", "📊 Gráficos", "📥 Reportes"])

# ============================================
# TAB 1: INGRESO DE DATOS
# ============================================
with tab1:
    st.header("📝 Registro de Nuevo Lote")
    
    with st.form("formulario_ingreso", clear_on_submit=True):
        col1, col2 = st.columns(2)
        
        with col1:
            st.subheader("📦 Información del Producto")
            
            # Estado
            estado = st.selectbox(
                "🏷️ Estado / Sección",
                ESTADOS,
                help="Estado actual del producto en bodega"
            )
            
            # Ubicación en bodega
            st.markdown("**📍 Ubicación en Bodega:**")
            col_ub1, col_ub2 = st.columns(2)
            with col_ub1:
                seccion = st.selectbox("Sección", SECCIONES_BODEGA, help="Sección de la bodega (A-J)")
            with col_ub2:
                fila = st.selectbox("Fila", FILAS_BODEGA, help="Número de fila (1-20)")
            
            st.markdown("---")
            
            # Selección de producto por código SAP o descripción
            st.markdown("**🔍 Buscar Producto:**")
            metodo_busqueda = st.radio(
                "Buscar por:",
                ["📋 Descripción", "🔢 Código SAP"],
                horizontal=True,
                label_visibility="collapsed"
            )
            
            if metodo_busqueda == "🔢 Código SAP":
                codigo_ingresado = st.number_input(
                    "Código SAP",
                    min_value=60000000,
                    max_value=69999999,
                    value=60000035,
                    step=1,
                    help="Ingresa el código SAP del producto"
                )
                if codigo_ingresado in PRODUCTOS_SAP:
                    descripcion = PRODUCTOS_SAP[codigo_ingresado]
                    codigo_sap = codigo_ingresado
                    st.success(f"✅ {descripcion}")
                else:
                    descripcion = ""
                    codigo_sap = codigo_ingresado
                    st.warning("⚠️ Código no encontrado en catálogo")
            else:
                producto_seleccionado = st.selectbox(
                    "Producto",
                    PRODUCTOS_LISTA,
                    help="Selecciona el producto del catálogo"
                )
                # Extraer código y descripción
                codigo_sap = int(producto_seleccionado.split(" - ")[0])
                descripcion = producto_seleccionado.split(" - ")[1]
            
            # Mostrar código SAP seleccionado
            st.markdown(f'<div class="seccion-info">🏷️ <b>Código SAP:</b> {codigo_sap}<br>📦 <b>Producto:</b> {descripcion}</div>', unsafe_allow_html=True)
            
            # Lote
            lote = st.text_input(
                "🔢 Número de Lote",
                placeholder="Ej: L5331, L5323C"
            )
            
            observaciones = st.text_area(
                "📝 Observaciones",
                placeholder="Notas adicionales...",
                height=80
            )
        
        with col2:
            st.subheader("📅 Fechas y Cantidades")
            
            col_f1, col_f2 = st.columns(2)
            with col_f1:
                fecha_elaboracion = st.date_input(
                    "📅 Fecha Elaboración",
                    value=datetime.now().date()
                )
            
            with col_f2:
                dias_vida = get_vida_util_dias(descripcion if descripcion else "DEFAULT")
                fecha_cad_sugerida = fecha_elaboracion + timedelta(days=dias_vida)
                fecha_caducidad = st.date_input(
                    "📅 Fecha Caducidad",
                    value=fecha_cad_sugerida
                )
            
            # Preview de vida útil
            vida_util_preview = calcular_porcentaje_vida_util(fecha_elaboracion, fecha_caducidad)
            semaforo_preview = semaforo_vida_util(vida_util_preview)
            
            if vida_util_preview is not None:
                if vida_util_preview >= 50:
                    st.success(f"{semaforo_preview} Vida útil: **{vida_util_preview:.1f}%** - Óptimo")
                elif vida_util_preview >= 30:
                    st.warning(f"{semaforo_preview} Vida útil: **{vida_util_preview:.1f}%** - Alerta")
                else:
                    st.error(f"{semaforo_preview} Vida útil: **{vida_util_preview:.1f}%** - Crítico")
            
            st.markdown("---")
            st.markdown("**📦 Cantidades:**")
            col_c1, col_c2, col_c3 = st.columns(3)
            
            with col_c1:
                pallets = st.number_input("🎯 Pallets", min_value=0, value=0, step=1)
            with col_c2:
                cajas = st.number_input("📦 Cajas", min_value=0, value=0, step=1)
            with col_c3:
                unidades = st.number_input("🔢 Unidades", min_value=0, value=0, step=1)
            
            # Resumen visual
            st.markdown("---")
            st.markdown("**📋 Resumen del Registro:**")
            st.markdown(f"""
            | Campo | Valor |
            |-------|-------|
            | **Ubicación** | Sección {seccion} - Fila {fila} |
            | **Código SAP** | {codigo_sap} |
            | **Producto** | {descripcion[:40]}... |
            | **Estado** | {estado} |
            """)
        
        # Botón de envío
        submitted = st.form_submit_button("✅ REGISTRAR PRODUCTO", type="primary", use_container_width=True)
        
        if submitted:
            if descripcion and lote:
                agregar_registro(
                    estado, seccion, fila, codigo_sap, descripcion, lote, 
                    fecha_elaboracion, fecha_caducidad, pallets, cajas, unidades, observaciones
                )
                st.success(f"✅ Lote **{lote}** - {descripcion} registrado en Sección {seccion}-{fila}!")
                st.balloons()
            else:
                st.error("⚠️ Completa al menos el Producto y el Lote")

# ============================================
# TAB 2: INVENTARIO ACTUAL
# ============================================
with tab2:
    st.header("📋 Inventario Actual")
    
    if st.session_state.inventario.empty:
        st.info("📭 No hay registros. Ve a **Ingreso de Datos** para agregar productos.")
    else:
        # Filtros en una fila
        st.subheader("🔍 Filtros")
        col_f1, col_f2, col_f3, col_f4, col_f5 = st.columns(5)
        
        with col_f1:
            filtro_estado = st.multiselect(
                "Estado",
                options=ESTADOS,
                default=st.session_state.inventario['ESTADO'].unique().tolist() if 'ESTADO' in st.session_state.inventario.columns else []
            )
        
        with col_f2:
            # Filtro por sección
            secciones_disponibles = st.session_state.inventario['SECCION'].unique().tolist() if 'SECCION' in st.session_state.inventario.columns else SECCIONES_BODEGA
            filtro_seccion = st.multiselect(
                "Sección",
                options=secciones_disponibles,
                default=secciones_disponibles
            )
        
        with col_f3:
            productos_unicos = st.session_state.inventario['DESCRIPCION'].unique().tolist() if 'DESCRIPCION' in st.session_state.inventario.columns else []
            filtro_producto = st.multiselect(
                "Producto",
                options=productos_unicos,
                default=productos_unicos[:10] if len(productos_unicos) > 10 else productos_unicos
            )
        
        with col_f4:
            filtro_semaforo = st.multiselect(
                "Vida Útil",
                options=["🔴 Crítico", "🟡 Alerta", "🟢 Óptimo"],
                default=["🔴 Crítico", "🟡 Alerta", "🟢 Óptimo"]
            )
        
        with col_f5:
            buscar_lote = st.text_input("🔎 Buscar Lote/SAP", "")
        
        # Aplicar filtros
        df_filtrado = st.session_state.inventario.copy()
        
        if filtro_estado and 'ESTADO' in df_filtrado.columns:
            df_filtrado = df_filtrado[df_filtrado['ESTADO'].isin(filtro_estado)]
        if filtro_seccion and 'SECCION' in df_filtrado.columns:
            df_filtrado = df_filtrado[df_filtrado['SECCION'].isin(filtro_seccion)]
        if filtro_producto and 'DESCRIPCION' in df_filtrado.columns:
            df_filtrado = df_filtrado[df_filtrado['DESCRIPCION'].isin(filtro_producto)]
        if buscar_lote:
            mask = df_filtrado['LOTE'].astype(str).str.contains(buscar_lote, case=False, na=False)
            if 'CODIGO_SAP' in df_filtrado.columns:
                mask = mask | df_filtrado['CODIGO_SAP'].astype(str).str.contains(buscar_lote, case=False, na=False)
            df_filtrado = df_filtrado[mask]
        
        # Filtro semáforo
        def filtrar_semaforo(row):
            vida = row.get('VIDA_UTIL_%')
            if pd.isna(vida) or vida is None:
                return True
            if vida < 30 and "🔴 Crítico" in filtro_semaforo:
                return True
            if 30 <= vida < 50 and "🟡 Alerta" in filtro_semaforo:
                return True
            if vida >= 50 and "🟢 Óptimo" in filtro_semaforo:
                return True
            return False
        
        if filtro_semaforo:
            df_filtrado = df_filtrado[df_filtrado.apply(filtrar_semaforo, axis=1)]
        
        # KPIs
        st.markdown("---")
        col_m1, col_m2, col_m3, col_m4, col_m5, col_m6 = st.columns(6)
        
        with col_m1:
            st.metric("📦 Cajas", f"{df_filtrado['CAJAS'].sum():,.0f}")
        with col_m2:
            st.metric("🎯 Pallets", f"{df_filtrado['PALLETS'].sum():,.0f}")
        with col_m3:
            st.metric("🔢 Unidades", f"{df_filtrado['UNIDADES'].sum():,.0f}")
        with col_m4:
            st.metric("📝 Registros", len(df_filtrado))
        with col_m5:
            criticos = len(df_filtrado[df_filtrado['VIDA_UTIL_%'] < 30]) if 'VIDA_UTIL_%' in df_filtrado.columns else 0
            st.metric("🔴 Críticos", criticos)
        with col_m6:
            secciones_ocupadas = df_filtrado['SECCION'].nunique() if 'SECCION' in df_filtrado.columns else 0
            st.metric("📍 Secciones", secciones_ocupadas)
        
        # Tabla de datos
        st.markdown("---")
        
        # Seleccionar columnas a mostrar (incluyendo nuevas)
        columnas_mostrar = ['ID', 'SECCION', 'FILA', 'ESTADO', 'CODIGO_SAP', 'DESCRIPCION', 'LOTE', 
                           'FECHA_ELABORACION', 'FECHA_CADUCIDAD', 'PALLETS', 'CAJAS', 
                           'UNIDADES', 'VIDA_UTIL_%', 'SEMAFORO']
        
        df_mostrar = df_filtrado[[c for c in columnas_mostrar if c in df_filtrado.columns]].copy()
        
        # Formatear y mostrar
        st.dataframe(
            df_mostrar.style.map(color_vida_util_bg, subset=['VIDA_UTIL_%'] if 'VIDA_UTIL_%' in df_mostrar.columns else [])
                          .map(color_estado_bg, subset=['ESTADO'] if 'ESTADO' in df_mostrar.columns else [])
                          .format({'VIDA_UTIL_%': '{:.1f}%'} if 'VIDA_UTIL_%' in df_mostrar.columns else {}),
            use_container_width=True,
            height=400
        )
        
        # Eliminar registro
        st.markdown("---")
        with st.expander("🗑️ Eliminar Registro"):
            if not df_filtrado.empty:
                opciones_eliminar = df_filtrado.apply(
                    lambda r: f"ID {r['ID']} - Sec:{r.get('SECCION', 'N/A')}-{r.get('FILA', 'N/A')} - {r.get('DESCRIPCION', 'N/A')[:30]} - Lote: {r['LOTE']}", axis=1
                ).tolist()
                
                seleccion = st.selectbox("Selecciona registro a eliminar:", opciones_eliminar)
                
                if seleccion:
                    id_seleccionado = int(seleccion.split(" - ")[0].replace("ID ", ""))
                    
                    col_del1, col_del2 = st.columns([1, 4])
                    with col_del1:
                        if st.button("🗑️ Eliminar", type="secondary"):
                            eliminar_registro(id_seleccionado)
                            st.warning(f"Registro eliminado")
                            st.rerun()

# ============================================
# TAB 3: GRÁFICOS Y ANÁLISIS
# ============================================
with tab3:
    st.header("📊 Análisis Visual")
    
    if st.session_state.inventario.empty:
        st.info("📭 No hay datos para graficar. Agrega registros primero.")
    else:
        df = st.session_state.inventario.copy()
        
        # Fila 1: Distribución por Estado y Producto
        col_g1, col_g2 = st.columns(2)
        
        with col_g1:
            st.subheader("📊 Distribución por Estado")
            df_estado = df.groupby('ESTADO')['CAJAS'].sum().reset_index()
            
            fig_pie = px.pie(
                df_estado,
                values='CAJAS',
                names='ESTADO',
                color='ESTADO',
                color_discrete_map={
                    'LIBRE UTILIZACIÓN': NUTRI_VERDE,
                    'RETENIDA': '#F44336',
                    'PRODUCCIÓN': '#FFC107',
                    'NOVEDADES': '#FF9800'
                },
                hole=0.4
            )
            fig_pie.update_layout(height=350, margin=dict(t=30, b=30))
            st.plotly_chart(fig_pie, use_container_width=True)
        
        with col_g2:
            st.subheader("📦 Top Productos por Cajas")
            df_producto = df.groupby('DESCRIPCION')['CAJAS'].sum().reset_index()
            df_producto = df_producto.sort_values('CAJAS', ascending=True).tail(10)
            
            fig_bar = px.bar(
                df_producto,
                x='CAJAS',
                y='DESCRIPCION',
                orientation='h',
                color='CAJAS',
                color_continuous_scale=[[0, NUTRI_AZUL_CLARO], [1, NUTRI_AZUL]]
            )
            fig_bar.update_layout(height=350, showlegend=False, margin=dict(t=30, b=30))
            st.plotly_chart(fig_bar, use_container_width=True)
        
        # Inventario por Sección
        st.markdown("---")
        st.subheader("📍 Inventario por Sección de Bodega")
        
        if 'SECCION' in df.columns and df['SECCION'].notna().any():
            df_seccion = df.groupby('SECCION').agg({
                'CAJAS': 'sum',
                'PALLETS': 'sum',
                'ID': 'count'
            }).rename(columns={'ID': 'LOTES'}).reset_index()
            
            fig_seccion = px.bar(
                df_seccion,
                x='SECCION',
                y='CAJAS',
                color='LOTES',
                text='CAJAS',
                color_continuous_scale=[[0, NUTRI_VERDE_CLARO], [1, NUTRI_VERDE]]
            )
            fig_seccion.update_traces(textposition='outside')
            fig_seccion.update_layout(height=300)
            st.plotly_chart(fig_seccion, use_container_width=True)
        
        # Semáforo de Vida Útil
        st.markdown("---")
        st.subheader("🚦 Semáforo de Vida Útil (FEFO)")
        
        col_s1, col_s2, col_s3 = st.columns(3)
        
        # Calcular grupos
        df_vida = df[df['VIDA_UTIL_%'].notna()].copy() if 'VIDA_UTIL_%' in df.columns else pd.DataFrame()
        criticos = df_vida[df_vida['VIDA_UTIL_%'] < 30] if not df_vida.empty else pd.DataFrame()
        alerta = df_vida[(df_vida['VIDA_UTIL_%'] >= 30) & (df_vida['VIDA_UTIL_%'] < 50)] if not df_vida.empty else pd.DataFrame()
        optimos = df_vida[df_vida['VIDA_UTIL_%'] >= 50] if not df_vida.empty else pd.DataFrame()
        
        with col_s1:
            st.markdown("### 🔴 Crítico (<30%)")
            st.markdown("**⚠️ Despacho Inmediato**")
            st.metric("Lotes", len(criticos))
            st.metric("Cajas", f"{criticos['CAJAS'].sum():,.0f}" if not criticos.empty else "0")
            if not criticos.empty:
                st.dataframe(
                    criticos[['LOTE', 'DESCRIPCION', 'VIDA_UTIL_%', 'CAJAS']].head(5),
                    hide_index=True,
                    height=150
                )
        
        with col_s2:
            st.markdown("### 🟡 Alerta (30-50%)")
            st.markdown("**⏰ Priorizar Despacho**")
            st.metric("Lotes", len(alerta))
            st.metric("Cajas", f"{alerta['CAJAS'].sum():,.0f}" if not alerta.empty else "0")
            if not alerta.empty:
                st.dataframe(
                    alerta[['LOTE', 'DESCRIPCION', 'VIDA_UTIL_%', 'CAJAS']].head(5),
                    hide_index=True,
                    height=150
                )
        
        with col_s3:
            st.markdown("### 🟢 Óptimo (>50%)")
            st.markdown("**✅ Estado Normal**")
            st.metric("Lotes", len(optimos))
            st.metric("Cajas", f"{optimos['CAJAS'].sum():,.0f}" if not optimos.empty else "0")
            if not optimos.empty:
                st.dataframe(
                    optimos[['LOTE', 'DESCRIPCION', 'VIDA_UTIL_%', 'CAJAS']].head(5),
                    hide_index=True,
                    height=150
                )
        
        # Gráfico de barras apiladas
        st.markdown("---")
        st.subheader("📈 Inventario por Producto y Estado")
        
        df_pivot = df.groupby(['DESCRIPCION', 'ESTADO'])['CAJAS'].sum().reset_index()
        
        fig_stacked = px.bar(
            df_pivot,
            x='DESCRIPCION',
            y='CAJAS',
            color='ESTADO',
            barmode='stack',
            color_discrete_map={
                'LIBRE UTILIZACIÓN': NUTRI_VERDE,
                'RETENIDA': '#F44336',
                'PRODUCCIÓN': '#FFC107',
                'NOVEDADES': '#FF9800'
            }
        )
        fig_stacked.update_layout(
            xaxis_tickangle=-45,
            height=400,
            legend=dict(orientation="h", yanchor="bottom", y=1.02),
            margin=dict(t=50, b=100)
        )
        st.plotly_chart(fig_stacked, use_container_width=True)
        
        # Estadísticas adicionales
        st.markdown("---")
        col_stat1, col_stat2 = st.columns(2)
        
        with col_stat1:
            st.subheader("📊 % por Estado")
            total_cajas = df['CAJAS'].sum()
            if total_cajas > 0:
                for estado in ESTADOS:
                    cajas_estado = df[df['ESTADO'] == estado]['CAJAS'].sum()
                    pct = (cajas_estado / total_cajas) * 100
                    color = {'LIBRE UTILIZACIÓN': '🟢', 'RETENIDA': '🔴', 'PRODUCCIÓN': '🟡', 'NOVEDADES': '🟠'}.get(estado, '⚪')
                    st.write(f"{color} **{estado}**: {pct:.1f}% ({cajas_estado:,.0f} cajas)")
        
        with col_stat2:
            st.subheader("📅 Próximos a Vencer")
            if 'FECHA_CADUCIDAD' in df.columns:
                df['FECHA_CADUCIDAD'] = pd.to_datetime(df['FECHA_CADUCIDAD'], errors='coerce')
                hoy = datetime.now()
                proximos_7_dias = df[df['FECHA_CADUCIDAD'] <= hoy + timedelta(days=7)]
                proximos_15_dias = df[(df['FECHA_CADUCIDAD'] > hoy + timedelta(days=7)) & 
                                      (df['FECHA_CADUCIDAD'] <= hoy + timedelta(days=15))]
                
                st.metric("⚠️ Vencen en 7 días", f"{len(proximos_7_dias)} lotes")
                st.metric("⏰ Vencen en 15 días", f"{len(proximos_15_dias)} lotes")

# ============================================
# TAB 4: REPORTES
# ============================================
with tab4:
    st.header("📥 Generación de Reportes")
    
    if st.session_state.inventario.empty:
        st.info("📭 No hay datos para exportar.")
    else:
        # Resumen
        st.subheader("📊 Resumen del Inventario")
        
        col_r1, col_r2 = st.columns(2)
        
        with col_r1:
            resumen = st.session_state.inventario.groupby('ESTADO').agg({
                'CAJAS': 'sum',
                'UNIDADES': 'sum',
                'PALLETS': 'sum',
                'ID': 'count'
            }).rename(columns={'ID': 'REGISTROS'}).reset_index()
            
            st.dataframe(resumen, use_container_width=True, hide_index=True)
        
        with col_r2:
            st.markdown("**📋 El reporte incluye:**")
            st.write("✅ Hoja con inventario completo")
            st.write("✅ Hojas separadas por estado")
            st.write("✅ Hoja de resumen ejecutivo")
            st.write("✅ Fechas y vida útil calculada")
        
        st.markdown("---")
        
        # Botón de descarga
        nombre_archivo = f"Inventario_Lacteos_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        
        col_btn1, col_btn2, col_btn3 = st.columns([1, 2, 1])
        with col_btn2:
            excel_data = exportar_excel()
            st.download_button(
                label="📥 DESCARGAR REPORTE EXCEL",
                data=excel_data,
                file_name=nombre_archivo,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                use_container_width=True
            )
        
        st.markdown("---")
        st.info(f"📁 Archivo: **{nombre_archivo}**")

# ============================================
# BARRA LATERAL
# ============================================
with st.sidebar:
    # Título con emoji en lugar de imagen
    st.markdown("## 🥛 NUTRI")
    st.markdown("### Lácteos San Antonio C.A.")
    st.markdown("**Sistema de Control de Inventario**")
    st.markdown("---")
    
    # Stats rápidos
    if not st.session_state.inventario.empty:
        st.markdown("### 📊 Resumen Rápido")
        
        total_registros = len(st.session_state.inventario)
        total_cajas = st.session_state.inventario['CAJAS'].sum()
        total_pallets = st.session_state.inventario['PALLETS'].sum()
        
        st.metric("📝 Registros", total_registros)
        st.metric("📦 Cajas", f"{total_cajas:,.0f}")
        st.metric("🎯 Pallets", f"{total_pallets:,.0f}")
        
        # Alertas
        if 'VIDA_UTIL_%' in st.session_state.inventario.columns:
            criticos = len(st.session_state.inventario[st.session_state.inventario['VIDA_UTIL_%'] < 30])
            if criticos > 0:
                st.error(f"🔴 {criticos} lotes CRÍTICOS")
        
        # Distribución por sección
        if 'SECCION' in st.session_state.inventario.columns:
            secciones_usadas = st.session_state.inventario['SECCION'].nunique()
            st.metric("📍 Secciones activas", secciones_usadas)
    else:
        st.info("📭 Sin datos aún")
    
    st.markdown("---")
    st.markdown("### 📖 Guía Rápida")
    st.markdown("""
    1. **📝 Ingreso**: Registrar lotes
    2. **📋 Inventario**: Ver/filtrar datos
    3. **📊 Gráficos**: Análisis visual
    4. **📥 Reportes**: Exportar Excel
    """)
    
    st.markdown("---")
    st.markdown(f"📅 {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    st.markdown("🧠 **GEM - Analítica de Datos**")
